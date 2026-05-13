#!/opt/venv/bin/python3
"""xlsx_inspect — paginated structural inspection of .xlsx workbooks.

Backed by openpyxl (read-only mode, constant memory) for cells, styles,
formulas, and defined names; drawings and merged-cell ranges are pulled
directly from the xlsx zip via stdlib so we don't have to re-load the
workbook in default mode.
"""

from __future__ import annotations

import argparse
import colorsys
import os
import re
import sys
import zipfile
from enum import Enum
from typing import Callable, Dict, List, NamedTuple, Optional, Tuple
from xml.etree import ElementTree as ET

import ooxml
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from utils import (
    TEXT_PREVIEW_LIMIT as VALUE_PREVIEW_LIMIT,
    ellipsize,
    format_size,
    pad,
    safe_output,
)

DEFAULT_LIMIT = 1000

EMU_PER_PIXEL = 9525
DEFAULT_COLUMN_PIXELS = 64
DEFAULT_ROW_PIXELS = 20

USAGE = (
    "xlsx_inspect <file> [--sheet NAME] [--range A1:Z50] "
    "[--grep PATTERN [--regex] [--meta]] [--names] [--limit N] [--offset N]"
)

HELP_TEXT = (
    "xlsx_inspect - Inspect .xlsx workbook structure\n"
    "\n"
    f"Usage: {USAGE}\n"
    "\n"
    "Arguments:\n"
    "  file              Path to .xlsx workbook (required)\n"
    "\n"
    "Options:\n"
    "  --sheet NAME      Show one sheet's cells (formula + cached value + format).\n"
    "  --range A1:Z50    Limit cells to a range. Requires --sheet.\n"
    "  --grep PATTERN    Search cell text across the workbook (or one sheet if --sheet).\n"
    "                    Substring match, case-insensitive. Use --grep '=' to list formulas.\n"
    "  --regex           Interpret --grep PATTERN as a regex.\n"
    "  --meta            Extend --grep to match (and emit) numFmt/font/fill tokens.\n"
    "                    e.g. --grep 'fill: FFFF' --meta finds yellow-ish highlights.\n"
    "  --names           List defined names with their ranges.\n"
    "  --limit N         Maximum rows/matches to print (default 1000).\n"
    "  --offset N        Skip first N rows/matches (default 0).\n"
    "\n"
    "Output (sheet view, one cell per line):\n"
    "  <address>  <formula or value>  [cached result]  numFmt: <fmt>  "
    "[font: <ARGB>]  [fill: <ARGB>]\n"
    "  Empty cells are skipped. Long strings are ellipsized."
)


class CellRange(NamedTuple):
    start_row: int
    start_col: int
    end_row: int
    end_col: int


def quote_string(value) -> str:
    if not isinstance(value, str):
        value = str(value)
    return '"' + ellipsize(value.replace("\n", "\\n"), VALUE_PREVIEW_LIMIT) + '"'


def format_scalar(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, str):
        return quote_string(value)
    iso = getattr(value, "isoformat", None)
    if callable(iso):
        return iso()[:10]
    return str(value)


_ASCII_A = ord("A")


# Convert between Excel column numbers (1-indexed: A=1, Z=26, AA=27)
# and their letter labels.
def col_letter(n: int) -> str:
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(_ASCII_A + r) + s
    return s


# Convert between Excel column letters to numbers
# this is the reciprocal function of col_letter
def col_number(letters: str) -> int:
    n = 0
    for ch in letters.upper():
        n = n * 26 + (ord(ch) - _ASCII_A + 1)
    return n


def parse_range(spec: str) -> CellRange:
    match = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", spec, re.IGNORECASE)
    if not match:
        raise ValueError(f'invalid --range: "{spec}" (expected e.g. A1:Z50)')
    start_col = col_number(match.group(1))
    start_row = int(match.group(2))
    end_col = col_number(match.group(3))
    end_row = int(match.group(4))
    if start_col > end_col or start_row > end_row:
        raise ValueError(f"invalid --range: {spec} (start must be <= end)")
    return CellRange(start_row, start_col, end_row, end_col)


# Black font and absent colors are treated as "default" — Excel writes
# FF000000 even for cells the user never styled. White font is meaningful
# (text on a colored fill), so it stays visible.
def is_default_font_color(argb: Optional[str]) -> bool:
    if not argb:
        return True
    return argb.upper() in ("FF000000", "00000000")


# Pure-white fills are de-facto "no highlight" in templates that fill
# every cell. Black fills (header banners) stay visible.
def is_default_fill_color(argb: Optional[str]) -> bool:
    if not argb:
        return True
    return argb.upper() in ("FFFFFFFF", "00FFFFFF")


def is_default_numfmt(numfmt: Optional[str]) -> bool:
    return not numfmt or numfmt == "General"


# Standard ECMA-376 §18.8.27 indexed palette. Slots 64 (system foreground)
# and 65 (system background) are sentinels we don't try to resolve.
INDEXED_PALETTE: Tuple[Optional[str], ...] = (
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF",
    "000000", "FFFFFF", "FF0000", "00FF00", "0000FF", "FFFF00", "FF00FF", "00FFFF",
    "800000", "008000", "000080", "808000", "800080", "008080", "C0C0C0", "808080",
    "9999FF", "993366", "FFFFCC", "CCFFFF", "660066", "FF8080", "0066CC", "CCCCFF",
    "000080", "FF00FF", "FFFF00", "00FFFF", "800080", "800000", "008080", "0000FF",
    "00CCFF", "CCFFFF", "CCFFCC", "FFFF99", "99CCFF", "FF99CC", "CC99FF", "FFCC99",
    "3366FF", "33CCCC", "99CC00", "FFCC00", "FF9900", "FF6600", "666699", "969696",
    "003366", "339966", "003300", "333300", "993300", "993366", "333399", "333333",
    None, None,
)


_THEME_DRAWING_NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main"
}

# XML element order in <a:clrScheme> vs. Excel's `theme` attribute index.
# Positions 0/1 and 2/3 swap: XML stores dk1, lt1, dk2, lt2; Excel uses
# 0=lt1, 1=dk1, 2=lt2, 3=dk2.
_THEME_INDEX_FOR_NAME = {
    "lt1": 0, "dk1": 1, "lt2": 2, "dk2": 3,
    "accent1": 4, "accent2": 5, "accent3": 6, "accent4": 7,
    "accent5": 8, "accent6": 9, "hlink": 10, "folHlink": 11,
}

# Canonical fallbacks for sysClr elements that omit lastClr.
_SYS_COLOR_DEFAULTS = {"windowText": "000000", "window": "FFFFFF"}


def _parse_theme_color(elem) -> Optional[str]:
    for child in elem:
        tag = child.tag.split("}", 1)[-1]
        if tag == "srgbClr":
            val = child.attrib.get("val")
            if isinstance(val, str):
                return val.upper()
        elif tag == "sysClr":
            last = child.attrib.get("lastClr")
            if isinstance(last, str):
                return last.upper()
            sys_name = child.attrib.get("val")
            fallback = _SYS_COLOR_DEFAULTS.get(sys_name) if sys_name else None
            if fallback:
                return fallback
    return None


def load_theme_palette(path: str) -> Dict[int, str]:
    """Map Excel `theme` index (0–11) to a 6-char hex string."""
    palette: Dict[int, str] = {}
    try:
        zip_file = zipfile.ZipFile(path)
    except zipfile.BadZipFile:
        return palette
    with zip_file:
        try:
            theme_stream = zip_file.open("xl/theme/theme1.xml")
        except KeyError:
            return palette
        with theme_stream:
            try:
                root = ET.parse(theme_stream).getroot()
            except ET.ParseError:
                return palette
        scheme = root.find(".//a:clrScheme", _THEME_DRAWING_NS)
        if scheme is None:
            return palette
        for child in scheme:
            name = child.tag.split("}", 1)[-1]
            idx = _THEME_INDEX_FOR_NAME.get(name)
            if idx is None:
                continue
            hex6 = _parse_theme_color(child)
            if hex6:
                palette[idx] = hex6
    return palette


def _apply_tint(hex6: str, tint: float) -> str:
    """Apply ECMA-376 §18.8.19 luminance tint to a 6-char hex color."""
    if not tint:
        return hex6
    r = int(hex6[0:2], 16) / 255.0
    g = int(hex6[2:4], 16) / 255.0
    b = int(hex6[4:6], 16) / 255.0
    h, l, s = colorsys.rgb_to_hls(r, g, b)
    l_new = l * (1.0 + tint) if tint < 0 else l + (1.0 - l) * tint
    l_new = max(0.0, min(1.0, l_new))
    r2, g2, b2 = colorsys.hls_to_rgb(h, l_new, s)

    def to_byte(v: float) -> int:
        return max(0, min(255, round(v * 255)))

    return f"{to_byte(r2):02X}{to_byte(g2):02X}{to_byte(b2):02X}"


def resolve_color(color, theme_palette: Dict[int, str]) -> Optional[str]:
    if color is None:
        return None
    try:
        ctype = getattr(color, "type", None)
        if ctype == "rgb":
            rgb = getattr(color, "rgb", None)
            return rgb if isinstance(rgb, str) else None
        if ctype == "theme":
            theme_idx = getattr(color, "theme", None)
            if not isinstance(theme_idx, int):
                return None
            base = theme_palette.get(theme_idx)
            if not base:
                return None
            tint = getattr(color, "tint", None) or 0.0
            return "FF" + _apply_tint(base, tint)
        if ctype == "indexed":
            idx = getattr(color, "indexed", None)
            if (
                not isinstance(idx, int)
                or idx < 0
                or idx >= len(INDEXED_PALETTE)
            ):
                return None
            entry = INDEXED_PALETTE[idx]
            return ("FF" + entry) if entry else None
    except Exception:
        return None
    return None


def font_argb(cell, theme_palette: Dict[int, str]) -> Optional[str]:
    font = getattr(cell, "font", None)
    if font is None:
        return None
    return resolve_color(getattr(font, "color", None), theme_palette)


def fill_argb(cell, theme_palette: Dict[int, str]) -> Optional[str]:
    fill = getattr(cell, "fill", None)
    if fill is None:
        return None
    if getattr(fill, "patternType", None) != "solid":
        return None
    return resolve_color(getattr(fill, "fgColor", None), theme_palette)


def describe_cell(cell, evaluated_value, theme_palette: Dict[int, str]) -> str:
    is_formula = cell.data_type == "f"
    formula_text = cell.value if is_formula else None

    if is_formula:
        if not isinstance(formula_text, str):
            return ""
        main = formula_text if formula_text.startswith("=") else f"={
            formula_text}"
    else:
        if cell.value is None:
            return ""
        main = format_scalar(cell.value)
        if not main:
            return ""

    if is_formula:
        result = format_scalar(evaluated_value) if evaluated_value is not None else ""
    else:
        result = ""

    numfmt = (
        ""
        if is_default_numfmt(cell.number_format)
        else f"numFmt: {cell.number_format}"
    )
    argb = font_argb(cell, theme_palette)
    color = "" if is_default_font_color(argb) else f"font: {argb.upper()}"
    fill_value = fill_argb(cell, theme_palette)
    fill = "" if is_default_fill_color(fill_value) else f"fill: {fill_value.upper()}"

    parts = [pad(cell.coordinate, 6), pad(ellipsize(main, 60), 32)]
    if result:
        parts.append(pad(result, 12))
    elif numfmt or color or fill:
        parts.append(pad("", 12))
    if numfmt:
        parts.append(numfmt)
    if color:
        parts.append(color)
    if fill:
        parts.append(fill)
    return "  ".join(parts).rstrip()


# ---------------------------------------------------------------------------
# xlsx-specific drawing & merge extraction.
#
# openpyxl in read_only mode skips xl/drawings/* and doesn't aggregate
# <mergeCells>, so we reach into the zip ourselves via ooxml helpers.
# Constant memory regardless of workbook size.
# ---------------------------------------------------------------------------

# xlsx-only namespaces (shared OOXML namespaces live in ooxml.NS).
_XL_NS = {
    **ooxml.NS,
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
}

_MERGE_TAG = f"{{{_XL_NS['main']}}}mergeCells"
_DRAWING_TAG = f"{{{_XL_NS['main']}}}drawing"
_ROW_TAG = f"{{{_XL_NS['main']}}}row"
_SHEET_DATA_TAG = f"{{{_XL_NS['main']}}}sheetData"


class AnchorKind(Enum):
    """How an OOXML drawing anchor is positioned on a worksheet.

    - TWO_CELL: pinned between two cells (start + end). Moves & resizes with rows/cols.
    - ONE_CELL: pinned at one cell with a fixed pixel extent (cx, cy in EMUs).
    - ABSOLUTE: pinned to absolute page coordinates, independent of any cell.
    """

    TWO_CELL = "xdr:twoCellAnchor"
    ONE_CELL = "xdr:oneCellAnchor"
    ABSOLUTE = "xdr:absoluteAnchor"


def _two_cell_anchor_range(anchor) -> Optional[str]:
    from_elem = anchor.find("xdr:from", _XL_NS)
    to_elem = anchor.find("xdr:to", _XL_NS)
    if from_elem is None or to_elem is None:
        return None
    start_col = int(from_elem.findtext("xdr:col", "0", _XL_NS)) + 1
    start_row = int(from_elem.findtext("xdr:row", "0", _XL_NS)) + 1
    end_col = int(to_elem.findtext("xdr:col", "0", _XL_NS)) + 1
    end_row = int(to_elem.findtext("xdr:row", "0", _XL_NS)) + 1
    return f"{col_letter(start_col)}{start_row}:{col_letter(end_col)}{end_row}"


def _one_cell_anchor_range(anchor) -> Optional[str]:
    from_elem = anchor.find("xdr:from", _XL_NS)
    if from_elem is None:
        return None
    start_col = int(from_elem.findtext("xdr:col", "0", _XL_NS)) + 1
    start_row = int(from_elem.findtext("xdr:row", "0", _XL_NS)) + 1
    ext = anchor.find("xdr:ext", _XL_NS)
    width_units = int(ext.attrib.get("cx", "0")) if ext is not None else 0
    height_units = int(ext.attrib.get("cy", "0")) if ext is not None else 0
    col_span = max(1, round(width_units / EMU_PER_PIXEL / DEFAULT_COLUMN_PIXELS))
    row_span = max(1, round(height_units / EMU_PER_PIXEL / DEFAULT_ROW_PIXELS))
    end_col = start_col + col_span - 1
    end_row = start_row + row_span - 1
    return f"{col_letter(start_col)}{start_row}:{col_letter(end_col)}{end_row} (estimate)"


def _absolute_anchor_range(_anchor) -> Optional[str]:
    return "absolute (floating, may overlap any cell)"


_ANCHOR_RESOLVERS: Tuple[Tuple[AnchorKind, Callable[[ET.Element], Optional[str]]], ...] = (
    (AnchorKind.TWO_CELL, _two_cell_anchor_range),
    (AnchorKind.ONE_CELL, _one_cell_anchor_range),
    (AnchorKind.ABSOLUTE, _absolute_anchor_range),
)


def _classify_anchor(
    zip_file: zipfile.ZipFile,
    anchor,
    drawing_rels: Dict[str, str],
    range_str: str,
) -> Optional[Tuple[str, str, str]]:
    graphic_frame = anchor.find("xdr:graphicFrame", _XL_NS)
    if graphic_frame is not None:
        chart_ref = graphic_frame.find("a:graphic/a:graphicData/c:chart", _XL_NS)
        title = "Chart"
        if chart_ref is not None:
            chart_relationship_id = chart_ref.attrib.get(ooxml.R_ID_ATTR)
            chart_path = (
                drawing_rels.get(chart_relationship_id) if chart_relationship_id else None
            )
            if chart_path:
                title = ooxml.parse_chart_title(zip_file, chart_path) or "Chart"
        return ("chart", title, range_str)
    if anchor.find("xdr:pic", _XL_NS) is not None:
        return ("image", "Image", range_str)
    return None


def _parse_drawing(
    zip_file: zipfile.ZipFile, drawing_path: str
) -> List[Tuple[str, str, str]]:
    tree = ooxml.read_xml(zip_file, drawing_path)
    if tree is None:
        return []
    drawing_rels = ooxml.parse_rels(zip_file, ooxml.rels_path_for(drawing_path))

    results: List[Tuple[str, str, str]] = []
    for kind, resolver in _ANCHOR_RESOLVERS:
        for anchor in tree.findall(kind.value, _XL_NS):
            cell_range = resolver(anchor)
            if not cell_range:
                continue
            classified = _classify_anchor(zip_file, anchor, drawing_rels, cell_range)
            if classified:
                results.append(classified)
    return results


SheetMeta = Dict[str, Dict[str, object]]


def _stream_sheet_meta(
    zip_file: zipfile.ZipFile, sheet_path: str
) -> Tuple[bool, Optional[str]]:
    """Stream-parse a sheet XML for merges + drawing rid.

    Sheet XML can be tens of MB on a busy workbook (sheetData dominates).
    `<mergeCells>` and `<drawing>` come after `<sheetData>`, so we
    iterparse and clear each `<row>` as we move past it to keep memory
    bounded.
    """
    merges = False
    drawing_relationship_id: Optional[str] = None
    try:
        sheet_stream = zip_file.open(sheet_path)
    except KeyError:
        return merges, drawing_relationship_id
    with sheet_stream:
        for _, elem in ET.iterparse(sheet_stream, events=("end",)):
            tag = elem.tag
            if tag == _MERGE_TAG:
                merges = len(elem) > 0
                elem.clear()
            elif tag == _DRAWING_TAG:
                drawing_relationship_id = elem.attrib.get(ooxml.R_ID_ATTR)
                elem.clear()
            elif tag == _ROW_TAG or tag == _SHEET_DATA_TAG:
                elem.clear()
    return merges, drawing_relationship_id


def parse_sheet_graphics(path: str) -> SheetMeta:
    """Map sheet name -> {"merges": bool, "graphics": [(kind, label, range_str)]}.

    Graphics = charts + images anchored to the sheet (from xl/drawings/*).
    """
    result: SheetMeta = {}
    try:
        zip_file = zipfile.ZipFile(path)
    except zipfile.BadZipFile:
        return result
    with zip_file:
        workbook_xml = ooxml.read_xml(zip_file, "xl/workbook.xml")
        if workbook_xml is None:
            return result
        workbook_rels = ooxml.parse_rels(zip_file, "xl/_rels/workbook.xml.rels")

        for sheet in workbook_xml.findall("main:sheets/main:sheet", _XL_NS):
            name = sheet.attrib.get("name")
            relationship_id = sheet.attrib.get(ooxml.R_ID_ATTR)
            if not name or not relationship_id:
                continue
            sheet_path = workbook_rels.get(relationship_id)
            if not sheet_path:
                continue

            merges, drawing_relationship_id = _stream_sheet_meta(zip_file, sheet_path)
            entry: Dict[str, object] = {"merges": merges, "graphics": []}
            if drawing_relationship_id:
                sheet_rels = ooxml.parse_rels(
                    zip_file, ooxml.rels_path_for(sheet_path)
                )
                drawing_path = sheet_rels.get(drawing_relationship_id)
                if drawing_path:
                    entry["graphics"] = _parse_drawing(zip_file, drawing_path)
            result[name] = entry
    return result


def sheet_graphics(meta: SheetMeta, sheet_name: str) -> List[Tuple[str, str, str]]:
    entry = meta.get(sheet_name) or {}
    graphics = entry.get("graphics") or []
    return graphics  # type: ignore[return-value]


def sheet_has_merges(meta: SheetMeta, sheet_name: str) -> bool:
    entry = meta.get(sheet_name) or {}
    return bool(entry.get("merges"))


def format_graphics(graphics) -> Optional[str]:
    if not graphics:
        return None
    lines = [f"[Graphics: {len(graphics)}]"]
    for kind, label, range_str in graphics:
        if kind == "chart" and label != "Chart":
            lines.append(f'- chart "{label}" at {range_str}')
        else:
            lines.append(f"- {kind} at {range_str}")
    return "\n".join(lines)


def graphics_summary(graphics) -> str:
    charts = sum(1 for kind, _, _ in graphics if kind == "chart")
    images = sum(1 for kind, _, _ in graphics if kind == "image")
    parts = []
    if charts:
        parts.append(f"charts: {charts}")
    if images:
        parts.append(f"images: {images}")
    return "  " + "  ".join(parts) if parts else ""


def find_sheet(workbook: Workbook, name: str) -> Worksheet:
    lower = name.lower()
    for sheet_name in workbook.sheetnames:
        if sheet_name.lower() == lower:
            return workbook[sheet_name]
    available = ", ".join(f'"{n}"' for n in workbook.sheetnames)
    raise ValueError(f'sheet not found: "{name}". Available: {available}')


def get_defined_names(workbook: Workbook):
    out = []
    defined_names = getattr(workbook, "defined_names", None)
    if defined_names is None:
        return out
    items = getattr(defined_names, "items", None)
    if callable(items):
        for name, definition in items():
            value = getattr(definition, "value", None)
            if value is None:
                destinations = getattr(definition, "destinations", None)
                if destinations:
                    value = ", ".join(f"{s}!{c}" for s, c in destinations)
            out.append((name, value or ""))
        return out
    inner = getattr(defined_names, "definedName", None)
    if inner:
        for definition in inner:
            out.append((getattr(definition, "name", "?"),
                       getattr(definition, "value", "") or ""))
    return out


def print_overview(workbook: Workbook, meta: SheetMeta) -> str:
    lines = [f"[Sheets: {len(workbook.sheetnames)}]"]
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        rows = worksheet.max_row or 0
        cols = worksheet.max_column or 0
        merges = "  merges: yes" if sheet_has_merges(meta, sheet_name) else ""
        graphics = graphics_summary(sheet_graphics(meta, sheet_name))
        state = ""
        if getattr(worksheet, "sheet_state", "visible") != "visible":
            state = f"  ({worksheet.sheet_state})"
        lines.append(
            f"- {pad(sheet_name, 24)} {pad(f'{rows}x{cols}', 10)}"
            f"{merges}{graphics}{state}"
        )
    names = get_defined_names(workbook)
    if names:
        lines.append("")
        lines.append(f"Defined names: {len(names)}")
        for name, value in names:
            lines.append(f"- {name} = {value}")
    return "\n".join(lines)


def print_names(workbook: Workbook) -> str:
    names = get_defined_names(workbook)
    if not names:
        return "[No defined names]"
    lines = [f"[Defined names: {len(names)}]"]
    for name, value in names:
        lines.append(f"- {name} = {value}")
    return "\n".join(lines)


def print_sheet(
    worksheet: Worksheet,
    evaluated_worksheet: Worksheet,
    graphics: List[Tuple[str, str, str]],
    range_spec: Optional[CellRange],
    offset: int,
    limit: int,
    theme_palette: Dict[int, str],
) -> str:
    total_rows = worksheet.max_row or 0
    if range_spec is not None:
        effective_start = max(range_spec.start_row, 1 + offset)
        last_row = min(range_spec.end_row, effective_start + limit - 1)
        start_col = range_spec.start_col
        end_col = range_spec.end_col
        end_row = range_spec.end_row
        range_label = (
            f"{col_letter(range_spec.start_col)}{range_spec.start_row}:"
            f"{col_letter(range_spec.end_col)}{range_spec.end_row}"
        )
    else:
        start_col = 1
        end_col = worksheet.max_column or 0
        end_row = total_rows
        effective_start = 1 + offset
        last_row = min(end_row, effective_start + limit - 1)
        range_label = f"Rows {effective_start}-{min(end_row, total_rows)}"

    lines = [f"[Sheet: {worksheet.title} | {range_label} | {total_rows} rows total]"]

    graphics_block = format_graphics(graphics)
    if graphics_block:
        lines.append(graphics_block)
        lines.append("")

    last_seen_row = effective_start - 1
    worksheet_iter = worksheet.iter_rows(
        min_row=effective_start, max_row=last_row,
        min_col=start_col, max_col=end_col,
    )
    evaluated_iter = evaluated_worksheet.iter_rows(
        min_row=effective_start, max_row=last_row,
        min_col=start_col, max_col=end_col,
    )
    for row_num, (row_cells, evaluated_cells) in enumerate(
        zip(worksheet_iter, evaluated_iter), start=effective_start
    ):
        any_in_row = False
        for cell, evaluated_cell in zip(row_cells, evaluated_cells):
            if cell.value is None:
                continue
            description = describe_cell(cell, evaluated_cell.value, theme_palette)
            if description:
                lines.append(description)
                any_in_row = True
        if any_in_row:
            last_seen_row = row_num

    if last_seen_row < end_row and last_seen_row < total_rows:
        lines.append("")
        lines.append(
            f"[Showing rows {effective_start}-{last_seen_row}. "
            f"Use --offset {last_seen_row} for the next page.]"
        )
    return "\n".join(lines)


def _cell_text(cell) -> Tuple[str, bool]:
    """Render a cell the same way --sheet mode does: formula source for
    formula cells, formatted scalar for value cells. Returns (text, is_formula);
    text is empty when the cell has nothing to print."""
    if cell.data_type == "f" and isinstance(cell.value, str):
        text = cell.value if cell.value.startswith("=") else f"={cell.value}"
        return text, True
    if cell.value is None:
        return "", False
    return format_scalar(cell.value), False


def print_grep(
    workbook: Workbook,
    evaluated_workbook: Workbook,
    pattern: str,
    *,
    regex: bool,
    sheet: Optional[str],
    range_spec: Optional[CellRange],
    offset: int,
    limit: int,
    theme_palette: Dict[int, str],
    include_meta: bool,
) -> str:
    if regex:
        try:
            compiled = re.compile(pattern, re.IGNORECASE)
        except re.error as exc:
            raise ValueError(f"invalid --grep regex: {exc}")

        def matches(text: str) -> bool:
            return bool(compiled.search(text))
    else:
        needle = pattern.lower()

        def matches(text: str) -> bool:
            return needle in text.lower()

    if sheet:
        sheet_names = [find_sheet(workbook, sheet).title]
    else:
        sheet_names = list(workbook.sheetnames)

    iter_kwargs: Dict[str, int] = {}
    if range_spec is not None:
        iter_kwargs = {
            "min_row": range_spec.start_row,
            "max_row": range_spec.end_row,
            "min_col": range_spec.start_col,
            "max_col": range_spec.end_col,
        }

    blocks: List[str] = []
    seen = 0
    shown = 0
    has_more = False

    for sheet_name in sheet_names:
        if has_more:
            break
        worksheet = workbook[sheet_name]
        evaluated_worksheet = evaluated_workbook[sheet_name]
        sheet_lines: List[str] = []

        for row_cells, evaluated_cells in zip(
            worksheet.iter_rows(**iter_kwargs),
            evaluated_worksheet.iter_rows(**iter_kwargs),
        ):
            if has_more:
                break
            for cell, evaluated_cell in zip(row_cells, evaluated_cells):
                text, is_formula = _cell_text(cell)
                if not text:
                    continue
                if include_meta:
                    description = describe_cell(
                        cell, evaluated_cell.value, theme_palette
                    )
                    if not description or not matches(description):
                        continue
                else:
                    description = ""
                    if not matches(text):
                        continue
                seen += 1
                if seen <= offset:
                    continue
                if shown >= limit:
                    has_more = True
                    break
                if include_meta:
                    sheet_lines.append(description)
                else:
                    if is_formula and evaluated_cell.value is not None:
                        result = format_scalar(evaluated_cell.value)
                    else:
                        result = ""
                    line = (
                        f"{pad(cell.coordinate, 6)}  "
                        f"{pad(ellipsize(text, 60), 32)}"
                    )
                    if result:
                        line += f"  {result}"
                    sheet_lines.append(line)
                shown += 1

        if sheet_lines:
            blocks.append(f"# {sheet_name}")
            blocks.extend(sheet_lines)
            blocks.append("")

    if seen == 0:
        kind = "regex" if regex else "pattern"
        return f"[No matches for {kind} {pattern!r}]"

    if blocks and blocks[-1] == "":
        blocks.pop()

    if has_more:
        header = f"[Matches: {shown} (more available)]"
    elif offset > 0:
        header = f"[Matches: {shown} (offset {offset})]"
    else:
        header = f"[Matches: {shown}]"

    out = [header, ""] + blocks
    if has_more:
        out.append("")
        out.append(f"[Use --offset {offset + shown} for the next page.]")
    return "\n".join(out)


def main() -> int:
    parser = argparse.ArgumentParser(
        prog="xlsx_inspect",
        usage=USAGE,
        add_help=False,
    )
    parser.add_argument("file", nargs="?")
    parser.add_argument("--sheet")
    parser.add_argument("--range", dest="range_spec")
    parser.add_argument("--grep")
    parser.add_argument("--regex", action="store_true")
    parser.add_argument("--meta", action="store_true")
    parser.add_argument("--names", action="store_true")
    parser.add_argument("--limit", type=int, default=DEFAULT_LIMIT)
    parser.add_argument("--offset", type=int, default=0)
    parser.add_argument("--help", "-h", action="store_true", dest="help_flag")

    args = parser.parse_args()

    if args.help_flag:
        sys.stdout.write(HELP_TEXT + "\n")
        return 0

    if not args.file:
        sys.stderr.write(f"Error: file is required\nUsage: {USAGE}\n")
        return 1

    if not os.path.isfile(args.file):
        sys.stderr.write(f"Error: file not found: {args.file}\n")
        return 1

    if args.limit < 1:
        sys.stderr.write("Error: --limit must be >= 1\n")
        return 1
    if args.offset < 0:
        sys.stderr.write("Error: --offset must be >= 0\n")
        return 1

    range_spec = None
    if args.range_spec:
        if not args.sheet:
            sys.stderr.write(
                f"Error: --range requires --sheet\nUsage: {USAGE}\n")
            return 1
        range_spec = parse_range(args.range_spec)

    workbook = openpyxl.load_workbook(args.file, data_only=False, read_only=True)
    needs_cached_values = bool(args.grep) or bool(args.sheet)
    evaluated_workbook = (
        openpyxl.load_workbook(args.file, data_only=True, read_only=True)
        if needs_cached_values
        else None
    )
    theme_palette = load_theme_palette(args.file)

    if args.grep is not None:
        body = print_grep(
            workbook, evaluated_workbook, args.grep,
            regex=args.regex, sheet=args.sheet, range_spec=range_spec,
            offset=args.offset, limit=args.limit,
            theme_palette=theme_palette, include_meta=args.meta,
        )
    elif args.names:
        body = print_names(workbook)
    elif args.sheet:
        sheet_meta = parse_sheet_graphics(args.file)
        worksheet = find_sheet(workbook, args.sheet)
        evaluated_worksheet = find_sheet(evaluated_workbook, args.sheet)
        body = print_sheet(
            worksheet, evaluated_worksheet,
            sheet_graphics(sheet_meta, worksheet.title),
            range_spec, args.offset, args.limit,
            theme_palette,
        )
    else:
        sheet_meta = parse_sheet_graphics(args.file)
        body = print_overview(workbook, sheet_meta)

    file_header = (
        f"[File: {os.path.basename(args.file)} | "
        f"{format_size(os.path.getsize(args.file))}]"
    )
    full = file_header + "\n" + body

    text, truncated = safe_output(full)
    sys.stdout.write(text)
    sys.stdout.write("\n")
    if truncated:
        sys.stdout.write(
            "[Output truncated; narrow with --range or paginate with "
            "--offset / --limit]\n"
        )
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except (ValueError, KeyError) as exc:
        sys.stderr.write(f"Error: {exc}\n")
        sys.exit(1)
    except Exception as exc:
        sys.stderr.write(f"Error: {type(exc).__name__}: {exc}\n")
        sys.exit(1)
